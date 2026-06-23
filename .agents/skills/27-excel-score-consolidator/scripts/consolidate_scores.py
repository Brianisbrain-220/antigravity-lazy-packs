#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 成績彙整 CLI 輔助腳本
讀取指定目錄中符合特定模式的班級資料夾，將其中的 CSV 評分檔合併並彙整成 Excel 活頁簿，
每個班級為一個工作表，包含「座號」、「姓名」、「各項作業分數」以及最後的「總平均」，並套用格式美化。
"""

import argparse
import csv
import glob
import json
import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def sanitize_sheet_title(title):
    """清理工作表名稱，使其符合 Excel 的長度（<=31字元）與非法字元限制。"""
    # 移除 Excel 工作表名稱不允許的字元: \ / ? * : [ ]
    clean_title = re.sub(r'[\\/\?\*\:\[\]]', '', title)
    return clean_title[:31]


def get_seat_sort_key(seat):
    """排序座號的鍵值函數：數字座號在前，非數字字串在後。"""
    if isinstance(seat, int):
        return (0, seat, "")
    try:
        return (0, int(seat), "")
    except (ValueError, TypeError):
        return (1, 0, str(seat))


def consolidate_scores(base_dir, class_pattern, output_path):
    """核心彙整邏輯：尋找班級資料夾、解析 CSV、產生 Excel。"""
    if not os.path.exists(base_dir):
        print(f"錯誤：指定的根目錄不存在：{base_dir}", file=sys.stderr)
        return None

    # 找出所有符合模式的班級資料夾
    class_dirs = []
    try:
        pattern_re = re.compile(class_pattern)
        for item in sorted(os.listdir(base_dir)):
            item_path = os.path.join(base_dir, item)
            if os.path.isdir(item_path) and pattern_re.search(item):
                class_dirs.append(item)
    except Exception as e:
        print(f"錯誤：班級名稱正則表達式解析失敗: {e}", file=sys.stderr)
        return None

    if not class_dirs:
        print(f"警告：在 {base_dir} 中沒有找到符合模式 '{class_pattern}' 的資料夾。", file=sys.stderr)
        return []

    wb = Workbook()
    # 移除預設建立的工作表
    wb.remove(wb.active)

    # ── 樣式定義 ────────────────────────────────────────────
    header_font = Font(name="微軟正黑體", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="微軟正黑體", size=11)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    avg_font = Font(name="微軟正黑體", bold=True, size=11, color="000000")
    avg_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    summary_log = []

    for class_dir in class_dirs:
        class_path = os.path.join(base_dir, class_dir)
        # 用前三個字作為工作表名稱，例如 "401運算思維" -> "401"
        sheet_title = sanitize_sheet_title(class_dir[:3])
        
        csv_files = sorted(glob.glob(os.path.join(class_path, "*.csv")))
        if not csv_files:
            continue

        assignments = {}
        for csv_file in csv_files:
            assignment_name = os.path.splitext(os.path.basename(csv_file))[0]
            scores = {}
            try:
                # 使用 utf-8-sig 自動處理有 BOM 的 UTF-8 檔案
                with open(csv_file, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        seat_raw = row.get("座號", "").strip()
                        name = row.get("姓名", "").strip()
                        score_raw = row.get("評定分數", "").strip()

                        if not seat_raw or not name:
                            continue

                        # 嘗試轉為整數，失敗則保留原始座號字串
                        try:
                            seat = int(seat_raw)
                        except ValueError:
                            seat = seat_raw

                        # 嘗試將分數轉為數值，非數值則保留原始字串
                        try:
                            score = int(score_raw)
                        except ValueError:
                            try:
                                score = float(score_raw)
                            except ValueError:
                                score = score_raw

                        scores[seat] = (name, score)
            except Exception as e:
                print(f"警告：讀取檔案 {csv_file} 失敗：{e}", file=sys.stderr)
                continue

            if scores:
                assignments[assignment_name] = scores

        if not assignments:
            continue

        # 彙整所有出現過的學生座號與姓名
        all_students = {}
        for asgn_scores in assignments.values():
            for seat, (name, _) in asgn_scores.items():
                all_students[seat] = name

        sorted_seats = sorted(all_students.keys(), key=get_seat_sort_key)
        assignment_names = list(assignments.keys())

        # 建立班級工作表
        ws = wb.create_sheet(title=sheet_title)
        headers = ["座號", "姓名"] + assignment_names + ["總平均"]

        # 寫入標題列並套用樣式
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 28

        # 寫入學生資料列
        for row_idx, seat in enumerate(sorted_seats, start=2):
            name = all_students[seat]

            # 寫入座號
            c_seat = ws.cell(row=row_idx, column=1, value=seat)
            c_seat.font = cell_font
            c_seat.alignment = cell_alignment
            c_seat.border = thin_border

            # 寫入姓名
            c_name = ws.cell(row=row_idx, column=2, value=name)
            c_name.font = cell_font
            c_name.alignment = Alignment(horizontal="left", vertical="center")
            c_name.border = thin_border

            numeric_scores = []
            # 寫入各項作業分數
            for asgn_idx, asgn_name in enumerate(assignment_names, start=3):
                asgn_data = assignments[asgn_name]
                if seat in asgn_data:
                    _, score = asgn_data[seat]
                    cell = ws.cell(row=row_idx, column=asgn_idx, value=score)
                    if isinstance(score, (int, float)):
                        numeric_scores.append(score)
                else:
                    cell = ws.cell(row=row_idx, column=asgn_idx, value="未繳")
                
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

            # 寫入總平均
            avg_col = len(headers)
            if numeric_scores:
                avg = round(sum(numeric_scores) / len(numeric_scores), 1)
            else:
                avg = "N/A"
            c_avg = ws.cell(row=row_idx, column=avg_col, value=avg)
            c_avg.font = avg_font
            c_avg.fill = avg_fill
            c_avg.alignment = cell_alignment
            c_avg.border = thin_border
            
            ws.row_dimensions[row_idx].height = 20

        # 自動調整每欄寬度
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # 中文字元長度計為 2，英文長度計為 1
                        char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                        max_len = max(max_len, char_len)
            
            # 換算 Excel 的欄名字元
            col_letter = ""
            temp = col_idx
            while temp > 0:
                modulo = (temp - 1) % 26
                col_letter = chr(65 + modulo) + col_letter
                temp = (temp - modulo) // 26
            
            ws.column_dimensions[col_letter].width = max_len + 4

        summary_log.append({
            "class": class_dir,
            "sheet_name": sheet_title,
            "students_count": len(sorted_seats),
            "assignments_count": len(assignment_names),
            "assignments": assignment_names
        })

    if summary_log:
        try:
            wb.save(output_path)
            return summary_log
        except Exception as e:
            print(f"錯誤：儲存 Excel 檔案失敗：{e}", file=sys.stderr)
            return None
    else:
        print("錯誤：未找到任何有效成績資料進行彙整。", file=sys.stderr)
        return []


def main():
    parser = argparse.ArgumentParser(description="多班級 CSV 成績單整理與 Excel 整合工具")
    subparsers = parser.add_subparsers(dest="command", required=True)

    p_consolidate = subparsers.add_parser("consolidate", help="彙整多班級成績並輸出 Excel 檔案")
    p_consolidate.add_argument("--base-dir", required=True, help="班級資料夾所在的根目錄路徑")
    p_consolidate.add_argument("--class-pattern", default=r"^4\d{2}", help="匹配班級資料夾名稱的正則表達式 (預設為 ^4\\d{2})")
    p_consolidate.add_argument("--output", required=True, help="輸出 Excel 檔案的完整路徑")
    p_consolidate.add_argument("--log", help="輸出彙整日誌的 JSON 檔案路徑")

    args = parser.parse_args()

    if args.command == "consolidate":
        result = consolidate_scores(args.base_dir, args.class_pattern, args.output)
        if result is None:
            sys.exit(1)
        
        # 成功訊息輸出至 stdout
        print(f"成功！已將成績資料整合並儲存至：{args.output}")
        for log_item in result:
            print(f"  - 班級: {log_item['class']} -> 工作表: {log_item['sheet_name']} ({log_item['students_count']} 位學生, {log_item['assignments_count']} 項作業)")
        
        # 將詳細紀錄輸出至指定的 JSON 檔案，以達 token 效率
        if args.log:
            try:
                with open(args.log, "w", encoding="utf-8") as f:
                    json.dump(result, f, indent=2, ensure_ascii=False)
                print(f"彙整日誌已寫入：{args.log}")
            except Exception as e:
                print(f"警告：寫入彙整日誌 JSON 失敗：{e}", file=sys.stderr)
    else:
        print(f"未知的指令：{args.command}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
