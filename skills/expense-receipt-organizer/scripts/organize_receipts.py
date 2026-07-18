import os
import sys
import re
import shutil
from datetime import datetime, timedelta
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# 預設為 True，跑 run 參數時為 False
DRY_RUN = True

# 1. 檔名合法化過濾
def clean_filename_part(s):
    if s is None:
        return ""
    s = re.sub(r'[\\/*?:"<>|]', "", str(s))
    s = re.sub(r'\s+', "_", s)
    s = re.sub(r'_+', "_", s)
    return s.strip("_")

# 2. 種類判定邏輯
def classify_expense(shop_name, item_name):
    shop = str(shop_name or "").lower()
    item = str(item_name or "").lower()
    
    # 交通
    if any(k in shop for k in ["租車", "レンタカー", "停車", "パーキング", "加油", "給油", "單軌電車", "モノレール", "uber", "交通", "車資"]):
        return "交通"
    if any(k in item for k in ["租車", "停車", "汽油", "儲值", "車資", "交通"]):
        return "交通"
        
    # 住宿
    if any(k in shop for k in ["住宿", "飯店", "酒店", "宿", "hotel", "airbnb"]):
        return "住宿"
    if any(k in item for k in ["住宿", "飯店", "房費"]):
        return "住宿"
        
    # 餐費
    if any(k in shop for k in ["燒肉", "牛", "大戶屋", "全家", "familymart", "lawson", "羅森", "7-11", "セブン", "maxvalu", "りうぼう", "超市", "波多馬", "ポーたま", "星巴克", "starbucks", "鹽屋", "塩屋", "產直市場", "玉泉洞", "餐", "喫茶", "食堂", "居酒屋"]):
        return "餐費"
    if any(k in item for k in ["餐", "飲", "食", "麵", "飯", "便當", "肉", "菜", "蛋", "奶", "茶", "咖啡", "軟糖", "餅", "甜甜圈", "蛋糕", "水"]):
        return "餐費"
        
    # 其他
    return "其他"

# 3. 讀取 Excel 的消費明細總表
def load_all_expenses(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.worksheets[0]
    
    all_records = []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], None
    
    header_idx = -1
    for idx, r in enumerate(rows):
        if any(isinstance(x, str) and ("消費日期" in x or "消費時間" in x) for x in r):
            header_idx = idx
            break
            
    if header_idx == -1:
        print("Error: Header not found in Excel")
        return [], None
        
    headers = [str(x).strip() if x is not None else "" for x in rows[header_idx]]
    date_col = -1
    shop_col = -1
    item_col = -1
    amount_col = -1
    
    for idx, h in enumerate(headers):
        if "日期" in h or "時間" in h:
            date_col = idx
        elif "店名" in h or "店家" in h or "名稱" in h:
            shop_col = idx
        elif "品項" in h or "項目" in h:
            item_col = idx
        elif "金額" in h or "價" in h or "JPY" in h:
            amount_col = idx
            
    for r_idx, r in enumerate(rows[header_idx+1:]):
        if r[date_col] is None or str(r[date_col]).strip() == "" or "總計" in str(r[date_col]):
            continue
            
        date_str = str(r[date_col]).strip()
        try:
            dt = datetime.strptime(date_str, "%Y/%m/%d %H:%M")
        except ValueError:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
                except ValueError:
                    continue
                    
        shop_name = r[shop_col]
        item_name = r[item_col] if item_col != -1 else ""
        amount = r[amount_col] if amount_col != -1 else 0
        category = classify_expense(shop_name, item_name)
        
        actual_row = r_idx + header_idx + 2
        
        all_records.append({
            "datetime": dt,
            "shop": shop_name,
            "item": item_name,
            "category": category,
            "amount": amount,
            "row": actual_row
        })
        
    print(f"Loaded {len(all_records)} expense records from Excel.")
    return all_records, sheet.title

# 4. 讀取圖片的 EXIF (動態讀取，以備無 report 檔案時使用)
def get_image_datetime(filepath):
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS
        with Image.open(filepath) as img:
            exif = img._getexif()
            if exif:
                for tag, value in exif.items():
                    decoded = TAGS.get(tag, tag)
                    if decoded == "DateTimeOriginal":
                        return datetime.strptime(value, "%Y:%m:%d %H:%M:%S")
    except Exception:
        pass
    return None

# 5. 載入 EXIF 資料 (支援讀取 report 或動態解析)
def load_exif_data(base_folder, report_path, files_list):
    image_dates = {}
    
    # 優先嘗試讀取 report 檔
    if report_path and os.path.exists(report_path):
        print(f"Reading EXIF from report file: {report_path}")
        try:
            with open(report_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or "|" not in line:
                        continue
                    parts = line.split("|")
                    if len(parts) >= 3:
                        filename = parts[0].strip()
                        base_name = os.path.basename(filename)
                        exif_part = parts[1].replace("EXIF:", "").strip()
                        mtime_part = parts[2].replace("mtime:", "").strip()
                        
                        exif_dt = None
                        if exif_part and exif_part != "None" and not exif_part.startswith("Error"):
                            exif_dt = datetime.strptime(exif_part, "%Y:%m:%d %H:%M:%S")
                                
                        mtime_dt = None
                        if mtime_part:
                            mtime_dt = datetime.strptime(mtime_part, "%Y-%m-%d %H:%M:%S")
                                
                        image_dates[base_name] = {"exif": exif_dt, "mtime": mtime_dt}
            return image_dates
        except Exception as e:
            print(f"Warning: Failed to read report file: {e}. Falling back to dynamic EXIF parsing...")
            
    # 若無 report 檔，則現場動態讀取圖片 EXIF
    print("No valid EXIF report file found. Dynamically extracting EXIF from images (this might take longer on cloud drives)...")
    try:
        from PIL import Image
    except ImportError:
        import subprocess
        print("Pillow not installed. Installing Pillow...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow"])
        
    for filename, filepath in files_list:
        if filename.lower().endswith(".jpg"):
            exif_dt = get_image_datetime(filepath)
            mtime_dt = datetime.fromtimestamp(os.path.getmtime(filepath))
            image_dates[filename] = {"exif": exif_dt, "mtime": mtime_dt}
            
    return image_dates

# 6. 遞迴掃描資料夾內所有單據檔案
def find_all_source_files(base_folder):
    all_files = []
    for root, dirs, files in os.walk(base_folder):
        for f in files:
            if f.endswith(".xlsx") or f == "desktop.ini" or f.endswith(".py") or f.endswith(".txt"):
                continue
            all_files.append((f, os.path.join(root, f)))
    print(f"Scanned {len(all_files)} single receipt files.")
    return all_files

# 7. 主邏輯
def organize_files(base_folder, excel_file, exif_report=None):
    expenses, sheet_title = load_all_expenses(excel_file)
    all_source_files = find_all_source_files(base_folder)
    image_dates = load_exif_data(base_folder, exif_report, all_source_files)
    
    move_actions = []
    excel_updates = {}
    
    for filename, filepath in all_source_files:
        base_name = os.path.basename(filename)
        
        # 移除可能已經存在的重命名首碼以重新分析
        clean_name = base_name
        clean_name = re.sub(r"^2026-\d\d-\d\d_[^_]+_[^_]+_\d+_", "", clean_name)
        clean_name = re.sub(r"^2026-\d\d-\d\d_[^_]+_", "", clean_name)
        
        target_date = None
        target_category = "其他"
        reason = ""
        matched_expense = None
        
        # A. PDF 檔案解析
        if clean_name.lower().endswith(".pdf"):
            if "uber" in clean_name.lower():
                m = re.search(r"7\s*(\d+)", clean_name)
                if m:
                    day = int(m.group(1))
                    target_date = datetime(2026, 7, day).strftime("%Y-%m-%d")
                    target_category = "交通"
                    reason = "PDF filename Uber"
            elif "住宿" in clean_name:
                m = re.search(r"7\s*(\d+)", clean_name)
                if m:
                    day = int(m.group(1))
                    target_date = datetime(2026, 7, day).strftime("%Y-%m-%d")
                    target_category = "住宿"
                    reason = "PDF filename 住宿"
            
            if not target_date:
                mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                target_date = mtime.strftime("%Y-%m-%d")
                reason = "PDF filename unmatched, fallback to file date"
                
        # B. JPG 檔案解析
        elif clean_name.lower().endswith(".jpg"):
            info = image_dates.get(clean_name, {"exif": None, "mtime": None})
            exif_dt = info["exif"]
            mtime_dt = info["mtime"]
            
            m_name = re.search(r"IMG202(\d)(\d\d)(\d\d)", clean_name)
            name_dt = None
            if m_name:
                year = 2020 + int(m_name.group(1))
                month = int(m_name.group(2))
                day = int(m_name.group(3))
                name_dt = datetime(year, month, day)
                
            base_dt = exif_dt if exif_dt else mtime_dt
            if base_dt is None:
                base_dt = datetime.fromtimestamp(os.path.getmtime(filepath))
                
            best_match = None
            min_diff = timedelta(hours=10)
            
            for r in expenses:
                diff = abs(r["datetime"] - base_dt)
                diff_tz1 = abs(r["datetime"] - (base_dt + timedelta(hours=1)))
                diff_tz2 = abs(r["datetime"] - (base_dt - timedelta(hours=1)))
                
                current_min = min(diff, diff_tz1, diff_tz2)
                if current_min < min_diff:
                    min_diff = current_min
                    best_match = r
                    
            if best_match and min_diff <= timedelta(minutes=30):
                target_date = best_match["datetime"].strftime("%Y-%m-%d")
                target_category = best_match["category"]
                matched_expense = best_match
                reason = f"Match Excel: {best_match['shop']} ({best_match['item']}), diff={min_diff}"
            else:
                if name_dt:
                    target_date = name_dt.strftime("%Y-%m-%d")
                    reason = "No Excel match, fallback to Name Date"
                else:
                    target_date = base_dt.strftime("%Y-%m-%d")
                    reason = f"No Excel match, fallback to file date"
                    
        else:
            mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
            target_date = mtime.strftime("%Y-%m-%d")
            reason = "Other file format"
            
        ext = os.path.splitext(clean_name)[1]
        base_no_ext = os.path.splitext(clean_name)[0]
        
        if matched_expense:
            shop_cleaned = clean_filename_part(matched_expense["shop"])
            amount_str = str(int(matched_expense["amount"] or 0))
            new_filename = f"{target_date}_{target_category}_{shop_cleaned}_{amount_str}_{base_no_ext}{ext}"
        else:
            new_filename = f"{target_date}_{target_category}_{base_no_ext}{ext}"
            
        move_actions.append({
            "src": filepath,
            "orig_filename": base_name,
            "new_filename": new_filename,
            "date": target_date,
            "category": target_category,
            "reason": reason,
            "matched_expense": matched_expense
        })
        
    print(f"\n==================== ACTIONS SUMMARY (DRY_RUN = {DRY_RUN}) ====================")
    for idx, act in enumerate(move_actions):
        target_dir = os.path.join(base_folder, act["date"], act["category"])
        target_path = os.path.join(target_dir, act["new_filename"])
        
        print(f"[{idx+1}] File: {act['orig_filename']} -> {act['new_filename']}")
        print(f"    -> Move to: {act['date']}/{act['category']}/")
        print(f"    -> Reason: {act['reason']}")
        
        relative_path = f"{act['date']}/{act['category']}/{act['new_filename']}"
        if act["matched_expense"]:
            row_idx = act["matched_expense"]["row"]
            excel_updates[row_idx] = relative_path
            
        if not DRY_RUN:
            if not os.path.exists(target_dir):
                os.makedirs(target_dir, exist_ok=True)
            try:
                shutil.move(act["src"], target_path)
                print(f"    [SUCCESS] Moved.")
            except Exception as e:
                print(f"    [FAILED] Error moving: {e}")
                
    if excel_updates and not DRY_RUN:
        print(f"\nWriting {len(excel_updates)} hyperlinks to Excel...")
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb[sheet_title]
            sheet.cell(row=1, column=8, value="單據憑證")
            
            for row_idx, rel_path in excel_updates.items():
                formula = f'=HYPERLINK("{rel_path}", "查看憑證")'
                sheet.cell(row=row_idx, column=8, value=formula)
                sheet.cell(row=row_idx, column=8).font = openpyxl.styles.Font(color="0000FF", underline="single")
                
            wb.save(excel_file)
            print("Excel hyperlinks updated successfully!")
        except Exception as e:
            print(f"Error updating Excel: {e}")
            
    print(f"\nProcessed {len(move_actions)} files.")

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "run":
        DRY_RUN = False
        
    # 可選參數：指定 EXIF 報告路徑作為加速用
    exif_report_path = None
    if os.path.exists(report_path):
        exif_report_path = report_path
        
    organize_files(folder, excel_path, exif_report=exif_report_path)
